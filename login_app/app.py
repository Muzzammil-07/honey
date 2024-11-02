from flask import Flask, render_template, request, redirect, url_for, flash
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Login credentials
my_username = {
    "muzzammil": "2007"
}

# Google Drive API setup
GOOGLE_DRIVE_FOLDER_ID = '10f6hxxJCEWgojQEdfwsa4KZL5OW6o-Jf'  # Replace with your Google Drive folder ID
credentials = service_account.Credentials.from_service_account_file('credentials.json')
drive_service = build('drive', 'v3', credentials=credentials)

# Function to save data to Excel and upload to Google Drive
def save_to_excel_and_upload(user_info):
    filename = "user_data.xlsx"
    
    # Create or load the workbook
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "user_data"
    
    # If it's a new file, add headers
    if sheet.max_row == 1:
        headers = ['name', 'branch', 'course', 'number']
        sheet.append(headers)
    
    # Append user data to Excel
    sheet.append([user_info['name'], user_info['branch'], user_info['course'], user_info['number']])
    workbook.save(filename)
    print("Data saved to Excel.")

    # Upload to Google Drive
    upload_to_google_drive(filename)

def upload_to_google_drive(filename):
    file_metadata = {
        'name': filename,
        'parents': [GOOGLE_DRIVE_FOLDER_ID]
    }
    media = MediaFileUpload(filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    print(f"File uploaded to Google Drive with ID: {file.get('id')}")

# Routes
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if username in my_username and my_username[username] == password:
            return redirect(url_for('user_info'))
        else:
            flash("Invalid username or password!")
    return render_template('index.html')

@app.route('/user_info', methods=['GET', 'POST'])
def user_info():
    if request.method == 'POST':
        # Collect user info from form
        user_data = {
            'name': request.form['name'],
            'branch': request.form['branch'],
            'course': request.form['course'],
            'number': request.form['number']
        }
        
        # Save data to Excel and upload to Google Drive
        save_to_excel_and_upload(user_data)
        flash("User data saved successfully!")
        return redirect(url_for('user_info'))
    
    return render_template('user_info.html')

if __name__ == '__main__':
    app.run(debug=True)
